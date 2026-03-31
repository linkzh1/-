# ui/device_list.py
# ========== Python 3.8 兼容性修复（必须放在文件最顶部）==========
import sys
import hashlib
import re
from dataclasses import dataclass
from typing import List, Dict, Optional
from datetime import datetime, timedelta
from collections import Counter, defaultdict
from pathlib import Path
import io
import json
import sqlite3
import socket
import platform
import os
from reportlab.platypus import Paragraph

if sys.version_info < (3, 9):
    _original_md5 = hashlib.md5


    def _patched_md5(data=b'', *, usedforsecurity=True):
        return _original_md5(data)


    hashlib.md5 = _patched_md5
    try:
        from reportlab.pdfbase import pdfdoc

        pdfdoc.md5 = _patched_md5
    except ImportError:
        pass

# ===============================================================
from flask import Blueprint, render_template, jsonify, request, send_file
from core.storage.database import Database
from core.vulnerability.scanner.engine import ScanEngine as VulnScannerEngine
from core.network.traffic_rate import TrafficMonitor


# ========== 问卷系统模块 ==========
@dataclass
class Question:
    id: str
    text: str
    type: str
    options: List[str]
    required: bool = False


class SurveyParser:
    """Markdown问卷解析器"""

    def __init__(self, md_content: str):
        self.content = md_content
        self.questions = []

    def parse(self) -> List[Question]:
        lines = self.content.split('\n')
        current_q = None
        i = 0

        while i < len(lines):
            line = lines[i].rstrip()

            if not line or line == '---':
                i += 1
                continue

            # 跳过一级/二级标题（欢迎语）
            if (line.startswith('# ') or line.startswith('## ')) and not line.startswith('### '):
                i += 1
                continue

            # 识别问题区块（### 开头）
            if line.startswith('### '):
                if current_q:
                    self.questions.append(current_q)

                title_raw = line.replace('### ', '').strip()
                icon = ''
                title = title_raw
                if title_raw and ord(title_raw[0]) > 127:
                    icon = title_raw[0]
                    title = title_raw[1:].strip()

                current_q = {
                    'icon': icon or '🔹',
                    'title': title,
                    'type': 'text',
                    'options': [],
                    'description': '',
                    'scale_labels': {'min': '1星', 'max': '5星'}
                }

                # 预读确定题型
                lookahead = []
                for j in range(1, 4):
                    if i + j < len(lines):
                        lookahead.append(lines[i + j].strip())

                for next_line in lookahead:
                    if not next_line:
                        continue
                    if next_line.startswith('- [ ]') or next_line.startswith('- [x]'):
                        current_q['type'] = 'multiple'
                        break
                    elif re.match(r'^\d+\.', next_line):
                        current_q['type'] = 'single'
                        break
                    elif '__SCALE__' in next_line:
                        current_q['type'] = 'scale'
                        if '1星' in next_line and '5星' in next_line:
                            parts = next_line.split('，')
                            for part in parts:
                                if '=' in part:
                                    kv = part.split('=')
                                    if len(kv) == 2:
                                        key, val = kv[0].strip(), kv[1].strip()
                                        if '1' in key:
                                            current_q['scale_labels']['min'] = val
                                        elif '5' in key:
                                            current_q['scale_labels']['max'] = val
                        i += 1
                        break
                    elif '__TEXT__' in next_line:
                        current_q['type'] = 'text'
                        i += 1
                        break

            # 收集选项
            elif line.startswith('- [ ]') or line.startswith('- [x]'):
                if current_q and current_q['type'] == 'multiple':
                    opt = line.replace('- [ ]', '').replace('- [x]', '').strip()
                    if opt:
                        current_q['options'].append(opt)

            elif re.match(r'^\d+\.', line) and current_q and current_q['type'] == 'single':
                opt = re.sub(r'^\d+\.\s*', '', line).strip()
                if opt:
                    current_q['options'].append(opt)

            # 收集描述
            elif line.startswith('- ') and current_q:
                desc_text = line.replace('- ', '').strip()
                if current_q['type'] == 'text' and not current_q['options']:
                    current_q['description'] = desc_text

            i += 1

        if current_q:
            self.questions.append(current_q)

        return [Question(
            id=f"q_{i}",
            text=q['text'],
            type=q['type'],
            options=q['options'],
            required=q.get('required', False)
        ) for i, q in enumerate(self.questions)]


class SurveyManager:
    """问卷数据管理器"""

    def __init__(self, db_path: str = "data/surveys.db"):
        self.db_path = db_path
        self._init_db()

    def _init_db(self):
        Path(self.db_path).parent.mkdir(exist_ok=True)
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS survey_responses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                survey_type TEXT NOT NULL,
                user_id TEXT,
                responses TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                completed BOOLEAN DEFAULT 0
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_survey_status (
                user_id TEXT PRIMARY KEY,
                pre_completed BOOLEAN DEFAULT 0,
                pre_completed_at TIMESTAMP,
                post_completed BOOLEAN DEFAULT 0,
                post_completed_at TIMESTAMP,
                skip_pre BOOLEAN DEFAULT 0
            )
        ''')

        conn.commit()
        conn.close()

    def load_survey(self, survey_type: str) -> str:
        file_map = {
            'pre': 'ui/surveys/notice_and_choice_pre_survey.md',
            'post': 'ui/surveys/notice_and_choice_post_survey.md'
        }
        filepath = file_map.get(survey_type)
        if not filepath or not Path(filepath).exists():
            raise FileNotFoundError(f"问卷文件不存在: {filepath}")
        with open(filepath, 'r', encoding='utf-8') as f:
            return f.read()

    def save_response(self, survey_type: str, responses: Dict, user_id: str = "anonymous"):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO survey_responses (survey_type, user_id, responses, completed)
            VALUES (?, ?, ?, 1)
        ''', (survey_type, user_id, json.dumps(responses, ensure_ascii=False)))

        now = datetime.now().isoformat()
        if survey_type == 'pre':
            cursor.execute('''
                INSERT OR REPLACE INTO user_survey_status 
                (user_id, pre_completed, pre_completed_at)
                VALUES (?, 1, ?)
            ''', (user_id, now))
        else:
            cursor.execute('''
                INSERT OR REPLACE INTO user_survey_status 
                (user_id, post_completed, post_completed_at)
                VALUES (?, 1, ?)
            ''', (user_id, now))

        conn.commit()
        conn.close()

    def check_status(self, user_id: str = "anonymous") -> Dict:
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT pre_completed, post_completed, skip_pre 
            FROM user_survey_status 
            WHERE user_id = ?
        ''', (user_id,))
        row = cursor.fetchone()
        conn.close()

        if row:
            return {'pre_completed': bool(row[0]), 'post_completed': bool(row[1]), 'skip_pre': bool(row[2])}
        return {'pre_completed': False, 'post_completed': False, 'skip_pre': False}

    def skip_pre_survey(self, user_id: str = "anonymous"):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO user_survey_status (user_id, skip_pre)
            VALUES (?, 1)
        ''', (user_id,))
        conn.commit()
        conn.close()


# 全局实例
survey_manager = SurveyManager()
traffic_monitor = TrafficMonitor(interval=5)
traffic_monitor.start_monitoring()
print("[+] 流量监控服务已启动")

# ========== 创建蓝图 ==========
device_bp = Blueprint('device', __name__)


# ========== 工具函数 ==========
def create_connection(db_path=None):
    """创建数据库连接"""
    conn = sqlite3.connect(db_path or 'data/devices.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS devices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ip TEXT, mac TEXT, vendor TEXT, device_type TEXT,
            status TEXT DEFAULT 'unknown',
            first_seen TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_scan TIMESTAMP, open_ports TEXT, services TEXT,
            risk_score REAL DEFAULT 0.0
        )
    ''')
    conn.commit()
    return conn


# ========== 流量监控API ==========
@device_bp.route('/api/traffic/current')
def get_traffic():
    try:
        stats = traffic_monitor.get_current_stats()
        return jsonify({
            'status': 'success',
            'upload_mbps': round(stats.get('upload_speed_bps', 0) / 1_000_000, 2),
            'download_mbps': round(stats.get('download_speed_bps', 0) / 1_000_000, 2),
            'connections': stats.get('total_connections', 0),
            'timestamp': datetime.now().strftime('%H:%M:%S')
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@device_bp.route('/api/traffic/history')
def get_traffic_history():
    try:
        count = request.args.get('count', 20, type=int)
        history = traffic_monitor.get_history(count=count)
        return jsonify({'status': 'success', 'data': history})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ========== 设备管理路由 ==========
@device_bp.route('/')
def index():
    """设备列表首页 - 集成问卷检查"""
    status = survey_manager.check_status()

    # 首次使用检查
    if not status['pre_completed'] and not status['skip_pre']:
        conn = create_connection('data/devices.db')
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM devices")
        count = cursor.fetchone()[0]
        conn.close()

        if count == 0:
            return render_template('survey_redirect.html', message="首次使用请先完成问卷调查")

    # 检查是否需要提示后调查
    show_post_survey_prompt = False
    if status['pre_completed'] and not status['post_completed']:
        conn = create_connection('data/devices.db')
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT COUNT(*) FROM active_vuln_results WHERE date(scan_time) = date('now')")
            if cursor.fetchone()[0] > 0:
                show_post_survey_prompt = True
        except:
            pass
        conn.close()

    # 获取设备列表
    conn = create_connection('data/devices.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM devices WHERE ip != 'unknown' AND mac != 'unknown'")
    devices = cursor.fetchall()
    conn.close()

    # 🆕 计算设备类型统计
    stats = {
        'total': len(devices),
        'router': 0,
        'camera': 0,
        'speaker': 0,
        'phone': 0,
        'other': 0
    }

    for d in devices:
        device_type = d[4] if len(d) > 4 else 'Unknown'
        dtype_lower = str(device_type).lower()

        if any(k in dtype_lower for k in ['路由', '网关', 'router', 'gateway']):
            stats['router'] += 1
        elif any(k in dtype_lower for k in ['摄像', 'camera', 'ipc']):
            stats['camera'] += 1
        elif any(k in dtype_lower for k in ['音箱', 'speaker', 'audio', 'sound']):
            stats['speaker'] += 1
        elif any(k in dtype_lower for k in ['手机', 'phone', 'mobile', '平板']):
            stats['phone'] += 1
        else:
            stats['other'] += 1

    device_list = []
    for d in devices:
        device_list.append({
            'id': d[0], 'ip': d[1], 'mac': d[2], 'vendor': d[3],
            'device_type': d[4] if len(d) > 4 else 'Unknown',
            'created_at': d[5] if len(d) > 5 else None
        })

    return render_template('index.html',
                           devices=device_list,
                           show_post_survey_prompt=show_post_survey_prompt,
                           stats=stats)


@device_bp.route('/device/<mac>')
def device_detail(mac):
    """设备详情页面"""
    conn = create_connection('data/devices.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM devices WHERE mac=?", (mac,))
    device = cursor.fetchone()

    if not device:
        conn.close()
        return "Device not found", 404

    device_dict = {
        'id': device[0], 'ip': device[1], 'mac': device[2],
        'vendor': device[3], 'device_type': device[4] if len(device) > 4 else 'Unknown',
        'created_at': device[5] if len(device) > 5 else None
    }

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS active_vuln_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_ip TEXT, device_mac TEXT, device_type TEXT,
            scan_time TEXT, vuln_type TEXT, severity TEXT,
            description TEXT, proof TEXT, fix_suggestion TEXT
        )
    ''')

    vulnerabilities = []
    try:
        cursor.execute('''
            SELECT vuln_type, severity, description, fix_suggestion, MAX(scan_time) as scan_time 
            FROM active_vuln_results 
            WHERE device_mac=? OR device_ip=?
            GROUP BY vuln_type, severity
            ORDER BY 
                CASE severity 
                    WHEN 'Critical' THEN 1 
                    WHEN 'CRITICAL' THEN 1
                    WHEN 'HIGH' THEN 2 
                    WHEN 'High' THEN 2 
                    WHEN 'MEDIUM' THEN 4 
                    WHEN 'Medium' THEN 4 
                    ELSE 6 
                END
        ''', (mac, device_dict['ip']))

        for row in cursor.fetchall():
            vulnerabilities.append({
                'type': row[0], 'severity': row[1], 'description': row[2],
                'proof': row[3], 'fix': row[4], 'scan_time': row[5]
            })
    except Exception as e:
        print(f"查询漏洞历史失败: {e}")

    conn.close()

    risk_score = 0
    if vulnerabilities:
        weights = {'CRITICAL': 10, 'HIGH': 7, 'MEDIUM': 4, 'LOW': 1}
        risk_score = min(sum(weights.get(v['severity'], 0) for v in vulnerabilities), 10)

    # ✅ 判断是否有扫描历史（基于是否查询到漏洞记录）
    has_scan_history = len(vulnerabilities) > 0

    return render_template('device_detail.html', device=device_dict,
                           vulnerabilities=vulnerabilities,
                           scan_history={'vuln_count': len(vulnerabilities),
                                         'risk_score': risk_score} if vulnerabilities else None,
                           has_scan_history=has_scan_history)


# ========== 扫描功能 ==========
@device_bp.route('/test/manual')
def manual_scan_page():
    return render_template('manual_scan.html')


@device_bp.route('/api/scan/manual', methods=['POST'])
def manual_scan():
    data = request.get_json()
    if not data or not data.get('ip'):
        return jsonify({'error': '缺少目标 IP 地址'}), 400

    target_device = {
        'ip': data.get('ip'),
        'mac': data.get('mac', '00:00:00:00:00:00'),
        'open_ports': data.get('ports', [80, 8080, 23, 2323, 8000, 1883]),
        'device_type': data.get('device_type', 'unknown'),
        'manufacturer': data.get('manufacturer', 'Test Device')
    }

    print(f"\n[手动扫描请求] 目标: {target_device['ip']}")
    engine = VulnScannerEngine()

    try:
        result = engine.scan_device(target_device)
        all_vulnerabilities = []

        for v in result.get('vulnerabilities', []):
            if isinstance(v, dict):
                vuln_name = v.get('cve_id') or v.get('title') or v.get('type') or 'Unknown Vulnerability'
                all_vulnerabilities.append({
                    'type': vuln_name, 'severity': v.get('severity', 'Medium'),
                    'description': v.get('description', v.get('title', '')),
                    'proof': v.get('proof', ''), 'fix': v.get('solution', v.get('fix', ''))
                })

        for v in result.get('auth_issues', []):
            if isinstance(v, dict):
                all_vulnerabilities.append({
                    'type': v.get('type', 'Auth Issue'), 'severity': v.get('severity', 'Medium'),
                    'description': v.get('details', str(v)),
                    'proof': v.get('proof', ''), 'fix': v.get('recommendation', v.get('fix', ''))
                })

        for v in result.get('protocol_issues', []):
            if isinstance(v, dict):
                all_vulnerabilities.append({
                    'type': v.get('protocol', 'Protocol Issue'), 'severity': v.get('severity', 'Medium'),
                    'description': v.get('issue', v.get('description', '')),
                    'proof': '', 'fix': v.get('recommendation', v.get('fix', ''))
                })

        weights = {'Critical': 10, 'HIGH': 7, 'High': 7, 'MEDIUM': 4, 'Medium': 4, 'LOW': 1, 'Low': 1}
        risk_score = min(sum(weights.get(v.get('severity', 'Low'), 1) for v in all_vulnerabilities), 10)

        if all_vulnerabilities:
            conn = create_connection('data/devices.db')
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS active_vuln_results (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    device_ip TEXT, device_mac TEXT, device_type TEXT,
                    scan_time TEXT, vuln_type TEXT, severity TEXT,
                    description TEXT, proof TEXT, fix_suggestion TEXT
                )
            ''')

            for vuln in all_vulnerabilities:
                cursor.execute('''
                    INSERT INTO active_vuln_results 
                    (device_ip, device_mac, device_type, scan_time, vuln_type, severity, description, proof, fix_suggestion)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    target_device['ip'], target_device['mac'], target_device['device_type'],
                    datetime.now().isoformat(), vuln.get('type', 'Unknown'),
                    vuln.get('severity', 'MEDIUM'), vuln.get('description', ''),
                    vuln.get('proof', ''), vuln.get('fix', '')
                ))
            conn.commit()
            conn.close()

            try:
                db = Database()
                for vuln in all_vulnerabilities:
                    db.add_vulnerability(
                        device_ip=target_device['ip'],
                        vuln={
                            'cve_id': vuln.get('type', 'UNKNOWN'),
                            'title': vuln.get('description', 'Unknown')[:50],
                            'description': vuln.get('description', ''),
                            'severity': vuln.get('severity', 'Medium'),
                            'cvss_score': weights.get(vuln.get('severity', 'Low'), 5.0),
                            'solution': vuln.get('fix', '')
                        }
                    )
            except Exception as e:
                print(f"[!] 保存到 vulnerabilities 表失败: {e}")

        return jsonify({
            'status': 'success', 'message': f"扫描完成，发现 {len(all_vulnerabilities)} 个漏洞",
            'target': target_device,
            'result': {
                'vuln_count': len(all_vulnerabilities), 'risk_score': risk_score,
                'vulnerabilities': all_vulnerabilities
            }
        })

    except Exception as e:
        import traceback
        error_msg = str(e)
        print(f"[手动扫描失败] {error_msg}")
        print(traceback.format_exc())
        return jsonify({'status': 'error', 'message': error_msg, 'traceback': traceback.format_exc()}), 500


@device_bp.route('/api/scan/status/<ip>')
def get_scan_status(ip):
    """获取设备扫描状态"""
    return jsonify({
        'status': 'completed',
        'progress': 100,
        'message': '扫描完成'
    })


@device_bp.route('/api/scan/network', methods=['POST'])
def scan_network():
    try:
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            local_ip = s.getsockname()[0]
            s.close()
            ip_parts = local_ip.split('.')
            network = f"{ip_parts[0]}.{ip_parts[1]}.{ip_parts[2]}.0/24"
        except:
            network = "192.168.1.0/24"

        print(f"\n[网络扫描请求] 网段: {network}")
        from core.network.arp_scanner import ARPScanner
        from core.network.device_identifier import DeviceIdentifier

        scanner = ARPScanner(network, timeout=2, max_workers=50)
        devices = scanner.scan()
        print(f"[+] ARP扫描完成，发现 {len(devices)} 个设备")

        identifier = DeviceIdentifier('data/oui.txt')
        conn = create_connection('data/devices.db')
        cursor = conn.cursor()
        new_devices, updated_devices = 0, 0

        for device in devices:
            vendor = identifier.identify(device.mac)
            device_type = "Unknown"

            if vendor:
                vendor_lower = vendor.lower()
                if any(x in vendor_lower for x in ['router', 'gateway', 'tp-link', 'huawei', 'xiaomi']):
                    device_type = "路由器/网关"
                elif any(x in vendor_lower for x in ['camera', 'ipc', 'dahua', 'hikvision']):
                    device_type = "摄像头"
                elif any(x in vendor_lower for x in ['phone', 'mobile', 'apple', 'samsung']):
                    device_type = "手机/平板"
                elif any(x in vendor_lower for x in ['computer', 'intel', 'dell', 'hp', 'lenovo']):
                    device_type = "计算机/服务器"

                # 统一归类：如果已经是"路由器"，改为"路由器/网关"
                if device_type == "路由器":
                    device_type = "路由器/网关"

            cursor.execute("SELECT id FROM devices WHERE mac=?", (device.mac,))
            if not cursor.fetchone():
                cursor.execute('''
                    INSERT INTO devices (ip, mac, vendor, device_type, status, first_seen, last_scan)
                    VALUES (?, ?, ?, ?, 'online', datetime('now'), datetime('now'))
                ''', (device.ip, device.mac, vendor, device_type))
                new_devices += 1
            else:
                cursor.execute('''
                    UPDATE devices SET ip=?, vendor=?, device_type=?, status='online', last_scan=datetime('now')
                    WHERE mac=?
                ''', (device.ip, vendor, device_type, device.mac))
                updated_devices += 1

        conn.commit()
        conn.close()

        return jsonify({
            'status': 'success',
            'message': f'扫描完成！发现 {len(devices)} 个设备（新增 {new_devices} 个，更新 {updated_devices} 个）',
            'devices_found': len(devices), 'new_devices': new_devices,
            'updated_devices': updated_devices, 'network': network
        })

    except PermissionError as e:
        return jsonify({'status': 'error', 'message': '权限不足，请以管理员身份运行程序'}), 403
    except Exception as e:
        import traceback
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ========== 问卷路由 ==========
@device_bp.route('/survey/<survey_type>')
def show_survey(survey_type):
    if survey_type not in ['pre', 'post']:
        return "Invalid survey type", 400

    try:
        md_content = survey_manager.load_survey(survey_type)
        parser = SurveyParser(md_content)
        questions = parser.parse()

        return render_template('survey.html',
                               survey_type=survey_type,
                               questions=questions,
                               title="使用前调查" if survey_type == 'pre' else "使用后调查")
    except Exception as e:
        print(f"加载问卷失败: {e}")
        return redirect('/')


@device_bp.route('/submit_survey/<survey_type>', methods=['POST'])
def submit_survey(survey_type):
    if survey_type not in ['pre', 'post']:
        return jsonify({'error': 'Invalid survey type'}), 400

    try:
        responses = {}
        for key, value in request.form.items():
            if key.startswith('q_'):
                responses[key] = value

        survey_manager.save_response(survey_type, responses)
        return render_template('survey_thanks.html',
                               survey_type=survey_type,
                               message="感谢您的参与！您的反馈将帮助我们改进产品。")
    except Exception as e:
        print(f"保存问卷失败: {e}")
        return jsonify({'error': str(e)}), 500


@device_bp.route('/api/survey/status')
def survey_status():
    return jsonify(survey_manager.check_status())


@device_bp.route('/survey/skip_pre', methods=['POST'])
def skip_pre_survey():
    survey_manager.skip_pre_survey()
    return jsonify({'status': 'success'})


@device_bp.route('/survey/redirect')
def survey_redirect():
    return render_template('survey_redirect.html')


# ========== 问卷管理后台 ==========
def calculate_survey_stats(responses_list):
    if not responses_list:
        return {'questions': {}, 'total': 0}

    question_stats = defaultdict(lambda: {'type': 'text', 'values': [], 'options_count': Counter()})

    for resp in responses_list:
        for q_key, value in resp.items():
            if isinstance(value, list):
                question_stats[q_key]['type'] = 'multiple'
                question_stats[q_key]['values'].extend(value)
                for v in value:
                    question_stats[q_key]['options_count'][str(v)] += 1
            elif isinstance(value, (int, float)):
                if 1 <= value <= 5:
                    question_stats[q_key]['type'] = 'scale'
                    question_stats[q_key]['values'].append(int(value))
                else:
                    question_stats[q_key]['values'].append(str(value))
                    question_stats[q_key]['options_count'][str(value)] += 1
            elif isinstance(value, str):
                try:
                    int_val = int(value)
                    if 1 <= int_val <= 5:
                        question_stats[q_key]['type'] = 'scale'
                        question_stats[q_key]['values'].append(int_val)
                    else:
                        question_stats[q_key]['type'] = 'single'
                        question_stats[q_key]['values'].append(value)
                        question_stats[q_key]['options_count'][value] += 1
                except ValueError:
                    question_stats[q_key]['type'] = 'single'
                    question_stats[q_key]['values'].append(value)
                    question_stats[q_key]['options_count'][value] += 1
            else:
                question_stats[q_key]['values'].append(str(value))
                question_stats[q_key]['options_count'][str(value)] += 1

    result = {}
    for q_key, data in question_stats.items():
        if data['type'] == 'scale':
            int_values = []
            for v in data['values']:
                try:
                    int_values.append(int(v))
                except (ValueError, TypeError):
                    continue

            result[q_key] = {
                'type': 'scale',
                'avg': round(sum(int_values) / len(int_values), 2) if int_values else 0,
                'distribution': dict(Counter(int_values)),
                'count': len(int_values)
            }
        else:
            top_options = data['options_count'].most_common(5)
            result[q_key] = {
                'type': data['type'],
                'top_options': top_options,
                'unique_count': len(data['options_count']),
                'count': len(data['values'])
            }

    return {'questions': result, 'total': len(responses_list)}


def get_survey_trend(cursor, days=7):
    trend = []
    for i in range(days - 1, -1, -1):
        date = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        cursor.execute("SELECT COUNT(*) FROM survey_responses WHERE date(created_at) = ?", (date,))
        count = cursor.fetchone()[0]
        trend.append({'date': date[-5:], 'count': count})
    return trend


@device_bp.route('/admin/surveys')
def admin_surveys():
    """问卷数据管理后台"""
    filter_type = request.args.get('type', 'all')
    date_range = request.args.get('date', '7')
    keyword = request.args.get('keyword', '')

    conn = sqlite3.connect('data/surveys.db')
    cursor = conn.cursor()

    where_clauses, params = [], []
    if filter_type != 'all':
        where_clauses.append("survey_type = ?")
        params.append(filter_type)
    if date_range != 'all':
        where_clauses.append(f"created_at >= datetime('now', '-{int(date_range)} days')")
    if keyword:
        where_clauses.append("responses LIKE ?")
        params.append(f'%{keyword}%')

    where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

    cursor.execute(f"""
        SELECT id, survey_type, responses, created_at 
        FROM survey_responses 
        {where_sql}
        ORDER BY created_at DESC
    """, params)

    surveys, all_responses = [], []
    for row in cursor.fetchall():
        responses = json.loads(row[2])
        surveys.append({
            'id': row[0], 'type': row[1],
            'type_label': '使用前' if row[1] == 'pre' else '使用后',
            'time': row[3], 'responses': responses
        })
        all_responses.append(responses)

    stats = calculate_survey_stats(all_responses)
    trend_data = get_survey_trend(cursor)
    conn.close()

    return render_template('admin_surveys.html',
                           surveys=surveys, stats=stats, trend_data=trend_data,
                           filter_type=filter_type, date_range=date_range,
                           keyword=keyword, total=len(surveys))


@device_bp.route('/admin/surveys/export/<format>')
def export_surveys(format):
    """导出问卷数据"""
    conn = sqlite3.connect('data/surveys.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM survey_responses ORDER BY created_at DESC")
    rows = cursor.fetchall()
    conn.close()

    if format == 'json':
        data = [{'id': row[0], 'type': row[1], 'user_id': row[2],
                 'responses': json.loads(row[3]), 'created_at': row[4]} for row in rows]
        return jsonify(data), 200, {
            'Content-Type': 'application/json',
            'Content-Disposition': f'attachment; filename=surveys_{datetime.now().strftime("%Y%m%d")}.json'
        }

    elif format == 'csv':
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['ID', 'Type', 'Created_At', 'Question', 'Answer'])
        for row in rows:
            responses = json.loads(row[3])
            for q_key, answer in responses.items():
                writer.writerow([row[0], row[1], row[4], q_key, json.dumps(answer, ensure_ascii=False)])
        output.seek(0)
        return output.getvalue(), 200, {
            'Content-Type': 'text/csv; charset=utf-8',
            'Content-Disposition': f'attachment; filename=surveys_{datetime.now().strftime("%Y%m%d")}.csv'
        }

    elif format == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment

            if not rows:
                return jsonify({'error': '没有数据可导出'}), 400

            wb = Workbook()
            ws = wb.active
            ws.title = "问卷数据"

            headers = ['ID', '问卷类型', '用户ID', '提交时间']
            sample_responses = json.loads(rows[0][3]) if rows else {}
            question_keys = sorted(sample_responses.keys())
            headers.extend([f'问题{i + 1}' for i in range(len(question_keys))])

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            for row_idx, row in enumerate(rows, 2):
                ws.cell(row=row_idx, column=1, value=row[0])
                ws.cell(row=row_idx, column=2, value='使用前' if row[1] == 'pre' else '使用后')
                ws.cell(row=row_idx, column=3, value=row[2] or 'anonymous')
                ws.cell(row=row_idx, column=4, value=row[4])

                try:
                    responses = json.loads(row[3])
                    for col_idx, q_key in enumerate(question_keys, 5):
                        answer = responses.get(q_key, '')
                        if isinstance(answer, list):
                            answer = ', '.join(str(a) for a in answer)
                        ws.cell(row=row_idx, column=col_idx, value=str(answer))
                except:
                    pass

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue(), 200, {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename=survey_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            }

        except ImportError:
            return jsonify({'error': '请先安装 openpyxl: pip install openpyxl'}), 500
        except Exception as e:
            import traceback
            return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@device_bp.route('/admin/surveys/delete/<int:id>', methods=['POST'])
def delete_survey(id):
    conn = sqlite3.connect('data/surveys.db')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM survey_responses WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ========== PDF报告 ==========
@device_bp.route('/api/report/pdf/<mac>')
def generate_pdf_report(mac):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        conn = create_connection('data/devices.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM devices WHERE mac=?", (mac,))
        device_row = cursor.fetchone()

        if not device_row:
            conn.close()
            return jsonify({'error': '设备不存在'}), 404

        device_data = {
            'ip': device_row[1], 'mac': device_row[2],
            'vendor': device_row[3] or 'Unknown',
            'device_type': device_row[4] or 'Unknown'
        }

        # ✅ 修改后的查询：按漏洞类型去重，保留最新记录，按严重程度排序
        cursor.execute('''
            SELECT vuln_type, severity, description, fix_suggestion, MAX(scan_time) as scan_time 
            FROM active_vuln_results 
            WHERE device_mac=? OR device_ip=?
            GROUP BY vuln_type
            ORDER BY 
                CASE severity 
                    WHEN 'Critical' THEN 1 
                    WHEN 'CRITICAL' THEN 1
                    WHEN 'HIGH' THEN 2 
                    WHEN 'High' THEN 2 
                    WHEN 'MEDIUM' THEN 3 
                    WHEN 'Medium' THEN 3 
                    WHEN 'LOW' THEN 4 
                    WHEN 'Low' THEN 4 
                    ELSE 5 
                END,
                scan_time DESC
        ''', (mac, device_data['ip']))

        vulnerabilities = []
        for row in cursor.fetchall():
            vulnerabilities.append({
                'type': row[0], 'severity': row[1],
                'description': row[2] or '暂无描述',
                'fix': row[3] or '建议联系厂商更新固件',
                'scan_time': row[4] or 'N/A'
            })
        conn.close()

        risk_score = 0
        if vulnerabilities:
            weights = {'Critical': 10, 'HIGH': 7, 'High': 7, 'MEDIUM': 4, 'Medium': 4, 'LOW': 1, 'Low': 1}
            risk_score = min(sum(weights.get(v.get('severity', 'Low'), 1) for v in vulnerabilities), 10)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        elements = []

        chinese_font = 'Helvetica'
        font_paths = [
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/simsun.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/System/Library/Fonts/PingFang.ttc"
        ]
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    font_name = "ChineseFont"
                    pdfmetrics.registerFont(TTFont(font_name, font_path))
                    chinese_font = font_name
                    break
                except:
                    continue

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=chinese_font, fontSize=24,
                                     alignment=1)
        normal_style = ParagraphStyle('Normal', parent=styles['BodyText'], fontName=chinese_font, fontSize=10)

        elements.append(Paragraph("IoT 设备安全评估报告", title_style))
        elements.append(Spacer(1, 0.5 * cm))
        elements.append(Paragraph(f"设备: {device_data['ip']} | {device_data['mac']}", normal_style))
        elements.append(Paragraph(f"风险评分: {risk_score}/10", normal_style))
        elements.append(Spacer(1, 1 * cm))

        if vulnerabilities:
            data = [['漏洞类型', '严重程度', '修复建议']]

            # 创建单元格段落样式（支持中文换行）
            from reportlab.platypus import Paragraph
            cell_style = ParagraphStyle(
                'CellStyle',
                parent=normal_style,
                fontName=chinese_font,
                fontSize=8,
                leading=12,  # 行间距
                wordWrap='CJK',
            )

            for v in vulnerabilities[:20]:
                # 用 Paragraph 包装，自动换行
                vuln_cell = Paragraph(v['type'][:40], cell_style)
                severity_cell = Paragraph(v['severity'], cell_style)
                fix_cell = Paragraph(v['fix'][:200], cell_style)  # 最多200字符

                data.append([vuln_cell, severity_cell, fix_cell])

            # 列宽：4.5cm, 2.5cm, 9cm（修复建议列最宽）
            table = Table(data, colWidths=[4.5 * cm, 2.5 * cm, 9 * cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), chinese_font),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        return send_file(buffer, mimetype='application/pdf',
                         as_attachment=True,
                         download_name=f"Report_{mac.replace(':', '-')}_{datetime.now().strftime('%Y%m%d')}.pdf")
    except Exception as e:
        return jsonify({'error': str(e)}), 500